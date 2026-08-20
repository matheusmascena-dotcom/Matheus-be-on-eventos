import base64, json, re, subprocess
from pathlib import Path
from openpyxl import load_workbook

BOOK=Path('data/Eventos - Be On _ 2026(5)_V3_Final.xlsx')
OUT=Path('assets/profile-matheus.jpg')
OUT.parent.mkdir(parents=True, exist_ok=True)
wb=load_workbook(BOOK, data_only=True)
ws=wb['Configurações do Site']
cfg={}
for r in ws.iter_rows(min_row=2, values_only=True):
    if r and r[0]: cfg[str(r[0]).strip()]=r[1]

def drive_id(v):
    if not v: return ''
    s=str(v)
    m=re.search(r'/file/d/([\w-]+)',s) or re.search(r'[?&]id=([\w-]+)',s)
    return m.group(1) if m else s
pid=drive_id(cfg.get('Foto Perfil'))
if not pid: raise SystemExit('Foto Perfil sem ID/link válido na planilha')
url=f'https://drive.usercontent.google.com/download?id={pid}&export=download&confirm=t'
subprocess.run(['curl','-L','--fail','--retry','3','-o',str(OUT),url],check=True)
if OUT.stat().st_size < 1000: raise SystemExit('Download da foto parece inválido')
# normalize to JPEG using Pillow
from PIL import Image
im=Image.open(OUT).convert('RGB')
im.thumbnail((720,960), Image.Resampling.LANCZOS)
im.save(OUT,'JPEG',quality=90,optimize=True,progressive=True)

site={
 'profile':'assets/profile-matheus.jpg',
 'instagram':str(cfg.get('Instagram') or ''),
 'whatsapp':str(cfg.get('WhatsApp') or ''),
 'coupon':str(cfg.get('Cupom') or ''),
 'discount':str(cfg.get('Desconto') or ''),
 'name':str(cfg.get('Nome Embaixador') or ''),
 'role':str(cfg.get('Cargo') or ''),
}
Path('site-config.json').write_text(json.dumps(site,ensure_ascii=False,indent=2),encoding='utf-8')

# Update only profile/contact/config references in HTML; leave event UI/features untouched.
old_img=re.compile(r'assets/(?:profile-photo|profile-matheus(?:-uploaded)?|matheus-profile)[^\"\']*\.(?:svg|png|jpg|jpeg)(?:\?[^\"\']*)?')
for p in [Path('index.html'),Path('event.html')]:
    if not p.exists(): continue
    s=p.read_text(encoding='utf-8')
    s=old_img.sub('assets/profile-matheus.jpg?v=sheet',s)
    if site['instagram']:
        s=re.sub(r'https://www\.instagram\.com/[^\"\']*',site['instagram'],s)
    if site['whatsapp']:
        s=re.sub(r'https://w\.app/[^\"\']*|https://wa\.me/[^\"\']*',site['whatsapp'],s)
    if site['coupon']:
        s=s.replace('MATHEUSMASCENA',site['coupon'])
    p.write_text(s,encoding='utf-8')
print('Sheet sync complete:', site)
